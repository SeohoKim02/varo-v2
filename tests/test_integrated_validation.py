"""Contracts for the button-only ten-sample integrated validation report."""
from __future__ import annotations

import copy
import io
import json
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from tests.streamlit_log_silencer import quiet_streamlit_test_logs

quiet_streamlit_test_logs()

from streamlit.testing.v1 import AppTest

from services.dqn_samples import DQN_SAMPLES, dqn_sample_path
from services.integrated_validation_service import (
    INTEGRATED_VALIDATION_VERSION,
    SUMMARY_COLUMNS,
    build_integrated_settings,
    build_sample_cache_key,
    clear_integrated_validation_cache,
    dqn_comparison_frame,
    error_result_frame,
    failed_sample_ids,
    integrated_detail_frame,
    integrated_excel_bytes,
    integrated_json_bytes,
    integrated_summary_frame,
    optimality_result_frame,
    run_integrated_validation,
    run_sample_validation,
    sample_catalog,
    sensitivity_result_frame,
    submission_summary_text,
)


def _minimal_sample(sample_id: str, status: str = "정상") -> dict:
    number = int(sample_id[-2:])
    return {
        "sample_id": sample_id,
        "sample_label": f"샘플 {number:02d}",
        "filename": f"sample_{number:02d}.xlsx",
        "metadata": {
            "store_count": 2, "dc_count": 1, "product_count": 3,
            "candidate_count": 2, "executable_candidate_count": 2,
            "load_status": "성공", "pipeline_status": "success",
            "source_unchanged": True,
        },
        "vhs": {"status": "정상", "average_vhs": 70.0, "total_expected_saving": 1000.0},
        "greedy": {"vhs_top1_match": True},
        "pareto": {"status": "보조 검증", "varo_top1_is_front": True},
        "dqn": {
            "original": {"training_status": "미학습", "target_type_count": 1, "target_dominant_ratio": 1.0},
            "balanced": {"training_status": "미학습", "target_type_count": 3, "target_dominant_ratio": 0.34},
            "numeric_features_preserved": True,
            "improved_vs_original": True,
        },
        "sensitivity": {"status": "계산 완료", "score": 90.0, "rating": "매우 안정"},
        "optimality": {
            "status": "정확 최적해", "search_status": "정확 최적해", "exact": True,
            "limited": False, "varo_gap_pct": 0.0, "target_pct": 100.0,
            "constraint_violation_count": 0, "unapplied_constraint_count": 0,
        },
        "overall_status": status,
        "status_reasons": [] if status == "정상" else ["테스트 실패"],
        "error_stage": "추천 pipeline 실행 중" if status == "실패" else None,
        "error_message": "boom" if status == "실패" else None,
        "excluded_calculations": ["후속 검증 전체"] if status == "실패" else [],
        "processing_seconds": 0.1,
        "cache_hit": False,
    }


def _minimal_result(status: str = "정상") -> dict:
    sample = _minimal_sample("sample_01", status)
    return {
        "version": INTEGRATED_VALIDATION_VERSION,
        "algorithm_version": "test",
        "settings": build_integrated_settings("quick", ["sample_01"]),
        "started_at": "2026-07-18T09:00:00",
        "completed_at": "2026-07-18T09:00:01",
        "samples": [sample],
        "summary": {
            "sample_count": 1,
            "normal_count": int(status == "정상"),
            "review_count": int(status == "확인 필요"),
            "failure_count": int(status == "실패"),
            "total_candidate_count": 2,
            "total_expected_saving": 1000.0,
            "average_vhs": 70.0,
            "average_sensitivity_score": 90.0,
            "average_target_pct": 100.0,
            "average_varo_gap_pct": 0.0,
            "total_processing_seconds": 1.0,
            "average_sample_seconds": 1.0,
            "cache_hit_count": 0,
        },
    }


class IntegratedValidationServiceTests(unittest.TestCase):
    def setUp(self):
        clear_integrated_validation_cache()

    def test_catalog_and_default_selection_are_exactly_ten(self):
        catalog = sample_catalog()
        self.assertEqual([row["sample_id"] for row in catalog], [f"sample_{number:02d}" for number in range(1, 11)])
        settings = build_integrated_settings()
        self.assertEqual(settings["scope"], "standard")
        self.assertEqual(len(settings["sample_ids"]), 10)
        self.assertFalse(settings["train_original"])
        self.assertFalse(settings["train_balanced"])

    def test_ten_samples_are_sequential_and_failure_does_not_stop_following_samples(self):
        seen: list[str] = []

        def fake_runner(sample, settings, progress_callback, force=False):
            sample_id = f"sample_{sample.number:02d}"
            seen.append(sample_id)
            return _minimal_sample(sample_id, "실패" if sample.number == 4 else "정상")

        with patch("services.integrated_validation_service.run_sample_validation", side_effect=fake_runner):
            result = run_integrated_validation(build_integrated_settings("quick"))
        self.assertEqual(seen, [f"sample_{number:02d}" for number in range(1, 11)])
        self.assertEqual(result["summary"]["sample_count"], 10)
        self.assertEqual(result["summary"]["failure_count"], 1)
        self.assertEqual(result["samples"][4]["sample_id"], "sample_05")

    def test_empty_selection_is_render_safe(self):
        result = run_integrated_validation(build_integrated_settings("quick", []))
        self.assertEqual(result["samples"], [])
        self.assertEqual(result["summary"]["sample_count"], 0)
        self.assertEqual(list(integrated_summary_frame(result).columns), SUMMARY_COLUMNS)

    def test_dqn_training_off_never_calls_training(self):
        settings = build_integrated_settings("quick", ["sample_01"])
        with patch("services.integrated_validation_service.train_dqn") as trainer:
            result = run_integrated_validation(settings)
        trainer.assert_not_called()
        self.assertEqual(result["summary"]["sample_count"], 1)
        self.assertNotEqual(result["samples"][0]["overall_status"], "실패")

    def test_full_scope_calls_only_selected_training_variant(self):
        settings = build_integrated_settings(
            "full", ["sample_01"], train_original=True, train_balanced=False,
        )
        fake_training = {
            "status": "정상", "final_status": "정상", "stability_status": "정상",
            "data_signature": "ignored", "prediction_distribution": {"할인": 2, "재고 이동": 2},
            "target_distribution": {"할인": 2, "재고 이동": 2}, "loss_history": [1.0, 0.2],
        }
        with (
            patch("services.integrated_validation_service.train_dqn", return_value=fake_training) as trainer,
            patch("services.integrated_validation_service.load_latest_dqn_result", return_value=None),
        ):
            result = run_integrated_validation(settings)
        self.assertEqual(trainer.call_count, 1)
        self.assertEqual(trainer.call_args.kwargs["training_mode"], "original")
        self.assertTrue(result["samples"][0]["dqn"]["numeric_features_preserved"])

    def test_standard_collects_sensitivity_gap_and_preserves_exact_status(self):
        settings = build_integrated_settings("standard", ["sample_01"])
        result = run_integrated_validation(settings)
        sample = result["samples"][0]
        self.assertGreater(sample["sensitivity"]["scenario_count"], 0)
        self.assertIn("exact", sample["optimality"])
        self.assertIn("limited", sample["optimality"])
        self.assertNotEqual(sample["sensitivity"]["status"], "계산 제외")
        self.assertTrue(sample["metadata"]["source_unchanged"])

    def test_source_workbook_hash_and_mtime_are_unchanged(self):
        path = dqn_sample_path(DQN_SAMPLES[0])
        before = (path.stat().st_size, path.stat().st_mtime_ns, path.read_bytes())
        result = run_integrated_validation(build_integrated_settings("quick", ["sample_01"]))
        after = (path.stat().st_size, path.stat().st_mtime_ns, path.read_bytes())
        self.assertEqual(before, after)
        self.assertTrue(result["samples"][0]["metadata"]["source_unchanged"])

    def test_cache_key_covers_file_hash_scope_and_training_flags(self):
        sample = DQN_SAMPLES[0]
        quick = build_integrated_settings("quick", ["sample_01"])
        standard = build_integrated_settings("standard", ["sample_01"])
        first = build_sample_cache_key(sample, "A", quick)
        self.assertEqual(first, build_sample_cache_key(sample, "A", copy.deepcopy(quick)))
        self.assertNotEqual(first, build_sample_cache_key(sample, "B", quick))
        self.assertNotEqual(first, build_sample_cache_key(sample, "A", standard))

    def test_success_cache_returns_copy_and_failed_selection_helper(self):
        settings = build_integrated_settings("quick", ["sample_01"])
        first = run_integrated_validation(settings)
        second = run_integrated_validation(settings)
        self.assertEqual(second["summary"]["cache_hit_count"], 1)
        second["samples"][0]["vhs"]["average_vhs"] = -1
        third = run_integrated_validation(settings)
        self.assertNotEqual(third["samples"][0]["vhs"]["average_vhs"], -1)
        self.assertEqual(failed_sample_ids(_minimal_result("실패")), ["sample_01"])

    def test_downloads_are_serializable_and_excel_has_required_sheets(self):
        result = _minimal_result()
        self.assertTrue(set(SUMMARY_COLUMNS).issubset(integrated_summary_frame(result).columns))
        for frame in (
            integrated_detail_frame(result), dqn_comparison_frame(result),
            sensitivity_result_frame(result), optimality_result_frame(result), error_result_frame(result),
        ):
            self.assertIsInstance(frame, pd.DataFrame)
        payload = integrated_json_bytes(result)
        self.assertEqual(json.loads(payload.decode("utf-8"))["version"], INTEGRATED_VALIDATION_VERSION)
        workbook = load_workbook(io.BytesIO(integrated_excel_bytes(result)), read_only=True)
        self.assertEqual(set(workbook.sheetnames), {
            "종합 요약", "VHS_Greedy", "DQN_Comparison", "Pareto_Sensitivity",
            "Optimality_Gap", "Errors", "Metadata",
        })
        self.assertIn("모델 성능 보장", submission_summary_text(result))


class IntegratedValidationAppTests(unittest.TestCase):
    def test_button_only_render_result_tabs_downloads_and_current_state_immutability(self):
        fake = _minimal_result()
        app = AppTest.from_file("app_v2.py", default_timeout=120)
        with patch("pages.validation.run_integrated_validation", return_value=fake) as runner:
            app.run()
            app.session_state["current_menu"] = "분석 및 검증"
            app.run()
            self.assertFalse(app.exception)
            self.assertIn("전체 샘플 통합 검증", [tab.label for tab in app.tabs])
            self.assertIn("통합 검증 실행", {button.label for button in app.button})
            self.assertIsNone(app.session_state["integrated_validation_result"])
            runner.assert_not_called()
            before = {
                key: copy.deepcopy(app.session_state[key])
                for key in ("varo_data", "varo_recommendations", "selected_route_id", "dqn_training_result", "sensitivity_result", "optimality_gap_result")
            }
            next(button for button in app.button if button.key == "run_integrated_validation").click().run()
        self.assertFalse(app.exception)
        runner.assert_called_once()
        for key, value in before.items():
            if key == "varo_data":
                self.assertEqual(value, app.session_state[key])
            else:
                self.assertEqual(value, app.session_state[key])
        self.assertEqual(app.session_state["integrated_validation_result"]["version"], INTEGRATED_VALIDATION_VERSION)
        metric_labels = {metric.label for metric in app.metric}
        self.assertTrue({"검증 샘플 수", "정상 수", "전체 예상 절감액", "전체 처리 시간"}.issubset(metric_labels))
        tab_labels = {tab.label for tab in app.tabs}
        self.assertTrue({
            "종합 요약", "VHS·Greedy", "DQN 원본·균형형", "Pareto·민감도",
            "최적성 Gap", "오류·제외 내역", "샘플별 상세",
        }.issubset(tab_labels))
        downloads = {button.label for button in app.get("download_button")}
        self.assertTrue({
            "전체 통합 요약 CSV", "전체 상세 결과 CSV", "DQN 비교 CSV", "민감도 결과 CSV",
            "최적성 Gap 결과 CSV", "오류 내역 CSV", "전체 JSON", "통합 Excel", "제출용 요약 TXT",
        }.issubset(downloads))

        app.session_state["current_menu"] = "홈"
        app.run()
        self.assertNotIn("통합 검증 실행", {button.label for button in app.button})


if __name__ == "__main__":
    unittest.main()
